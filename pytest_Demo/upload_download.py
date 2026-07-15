from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
import openpyxl


def update_excel_data(filePath, searchTerm, colName, new_value):
    book = openpyxl.load_workbook(filePath)
    sheet = book.active
    Dict = {}

    for i in range(1, sheet.max_column+ 1):
        if sheet.cell(row=1,column=i).value == colName:
            Dict["col"] = i

    for i in range(1, sheet.max_row+1):
        for j in range(1, sheet.max_column+1):
            if sheet.cell(row = i, column= j).value == searchTerm:
                Dict["row"] = i

    sheet.cell(row=Dict["row"], column =Dict["col"]).value = new_value
    book.save(filePath)